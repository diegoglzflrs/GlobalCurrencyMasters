import os
import requests
import json
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook, load_workbook

apikey = '7a86521fff0242e4b35bf70ee3696f80'

def menu():
    print("""###Global###Currency###Masters###
###           ###                ###              ###
###           ###                ###              ###
###           ###                ###              ###
1. Consultar tipos de monedas
2. Convertir fuente a blanco
3. Consultar tasas de conversiones recientes
4. Comparar valor de monedas extranjeras
5. Salir""")

def consultarMonedas():
    url = f'https://api.currencyfreaks.com/v2.0/supported-currencies'
    buscar = input("Ingrese el código de moneda que desea buscar: ").upper()
    respuesta = requests.get(url)
    if respuesta.status_code == 200:
        datos = respuesta.json()
        texto_json = json.dumps(datos, indent=2)
        with open("consulta.txt", "w") as archivo:
            archivo.write(texto_json)
        for codigo, nombre in datos.items(): 
            if buscar in codigo or buscar in nombre:
                print("Si existe")
            else:
                print("No existe")
    else:
        print(f"Error: No se pudo encontrar la información.")
    print("\n")

def convertirMonedas():
    try:
        fuente = input("Ingrese el código de moneda fuente: ").upper()
        blanco = input("Ingrese el código de moneda de destino: ").upper()
        cant = float(input("Ingrese la cantidad a convertir: "))
        url = f"https://api.currencyfreaks.com/v2.0/rates/latest?apikey={apikey}&symbols={blanco}&base={fuente}"
        respuesta = requests.get(url)
        if respuesta.status_code == 200:
            datos = respuesta.json()
            rate = float(datos['rates'][blanco])
            convertido = cant * rate
            if convertido:
                print(f'{cant} {fuente} es igual a {convertido} {blanco}')
                fecha_hora = datetime.now().isoformat() 
                with open("conversion.txt", "a") as archivo:
                    archivo.write(json.dumps({
                        "fecha": fecha_hora,
                        "fuente": fuente,
                        "destino": blanco,
                        "tasa": rate
                    }) + "\n")

                    excel = "conversiones.xlsx"
                    if os.path.exists(excel):
                        libro = load_workbook(excel)
                        hoja = libro.active
                    else:
                        libro = Workbook()
                        hoja = libro.active
                        hoja.append(["Fecha", "Moneda Fuente", "Moneda Destino", "Cantidad", "Tasa", "Resultado"])
                    hoja.append([fecha_hora, fuente, blanco, cant, rate, convertido])
                    libro.save(excel)
            else:
                print('Error: No se encontraron datos de conversión en la respuesta del API')
        else:
            print(f'Error: Codigo de estado: {respuesta.status_code}')
    except ValueError as e:
        print("Error: ", e)
    print("\n")

def consultarTasas():
    graficarTasas()

def graficarTasas():
    fechas = []
    tasas = []

    with open("conversion.txt", "r") as archivo:
        for linea in archivo:
            try:
                dato = json.loads(linea.strip())
                fechas.append(dato['fecha'])
                tasas.append(float(dato['tasa']))
            except Exception as e:
                continue

    if len(tasas) < 2:
        print("No hay suficientes datos para graficar.")
        return

    plt.figure(figsize=(12, 6))
    plt.plot(fechas, tasas, marker='o', color='blue')
    plt.xticks(rotation=45)
    plt.title("Evolución de la tasa de conversión")
    plt.xlabel("Fecha y Hora")
    plt.ylabel("Tasa de cambio")
    plt.tight_layout()
    plt.grid(True)
    plt.show()

def graficarComparacionMonedas():
    monedas = ['USD', 'EUR', 'MXN', 'JPY', 'ARS']
    colores = ['orange', 'red', 'darkorange', 'magenta', 'skyblue']
    valores = {}

    url = f"https://api.currencyfreaks.com/v2.0/rates/latest?apikey={apikey}"
    respuesta = requests.get(url)

    if respuesta.status_code != 200:
        print(f"Error al obtener datos. Código HTTP: {respuesta.status_code}")
        return

    datos = respuesta.json()
    rates = datos.get('rates', {})

    try:
        mxn_rate = float(rates['MXN'])  # Cuánto vale 1 USD en MXN
    except KeyError:
        print("No se encontró la tasa de MXN.")
        return

    for moneda in monedas:
        try:
            if moneda == 'MXN':
                valores[moneda] = 1.0
            else:
                tasa = float(rates[moneda])  # Cuánto vale 1 USD en esa moneda
                valores[moneda] = mxn_rate / tasa  # Valor relativo en MXN
        except KeyError:
            print(f"No se encontró la tasa para {moneda}")
        except ValueError:
            print(f"Tasa inválida para {moneda}")

    if len(valores) < 2:
        print("No hay suficientes tasas válidas para graficar.")
        return

    cantidades = list(range(1, 101))
    plt.figure(figsize=(12, 6))

    for i, moneda in enumerate(monedas):
        if moneda in valores:
            conversiones = [cantidad * valores[moneda] for cantidad in cantidades]
            plt.plot(cantidades, conversiones, label=moneda, color=colores[i])

    plt.title("Comparación del valor de monedas extranjeras en MXN (1–100 unidades)")
    plt.xlabel("Cantidad de moneda extranjera")
    plt.ylabel("Valor en pesos mexicanos (MXN)")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.show()

def salir():
    print("Adios!")
    return False

def main():
    opciones ={
        1: consultarMonedas,
        2: convertirMonedas,
        3: consultarTasas,
        4: graficarComparacionMonedas,
        5: salir
    }
    menu()
    banderaMenu = True
    while banderaMenu:
        try:
            op = int(input("Elige el número de la opción que desea realizar: "))
            elegida = opciones.get(op)
            if elegida:
                resultado = elegida()
                if resultado is False:
                    banderaMenu = False
            else:
                print("Error: Opción inválida.")
        except ValueError:
            print("Error: Debes ingresar un número.")

main()
